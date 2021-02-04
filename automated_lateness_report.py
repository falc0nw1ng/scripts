from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import os

## login information and login in to ClearCare
def date_inputs():
    print('Enter your starting date in the following format mm/dd/yyyy')
    first_date = input()
    print('Enter the ending date in the following format mm/dd/yyyy')
    second_date = input()
    return [first_date, second_date]

date_ranges = date_inputs()
options = webdriver.ChromeOptions()
cur = os.getcwd()
download_PATH = os.path.dirname(os.path.dirname(os.path.dirname(cur))) + '\\input'
prefs = {"profile.default_content_settings.popups": 0,
             "download.default_directory": download_PATH,
             "directory_upgrade": True}
options.add_experimental_option("prefs", prefs)

# either replace \ with \\ or use r''
driver_PATH = os.getcwd() + '\\chromedriver.exe'
#r"C:\Users\Codee Lau\Desktop\chromedriver\chromedriver.exe"
driver = webdriver.Chrome(driver_PATH, options=options)
#login_info
login_file = open('..\..\..\login_info.txt', 'r',)
login_info = []
for i in login_file.read().splitlines():
    login_info.append(i)

login_file.close()

driver.get("https://proofofcare.clearcareonline.ca/login/?next=/dashboard/")
enter_username = driver.find_element_by_id("id_username")
enter_username.send_keys(login_info[0])

enter_password = driver.find_element_by_id('id_password')
enter_password.send_keys(login_info[1])
enter_password.send_keys(Keys.RETURN)

driver.implicitly_wait(10)
reports_page = 'https://proofofcare.clearcareonline.ca/reports/custom/extreme/#/saved/carelogs-by-caregiver/90021'
driver.get(reports_page)
driver.implicitly_wait(10)

date1 = driver.find_element_by_xpath("(//input[@class='ng-pristine ng-untouched ng-valid hasDatepicker ng-not-empty ng-valid-required'])[1]")
date1.clear()
date1.send_keys(date_ranges[0])
date1.send_keys(Keys.ESCAPE)
date2 = driver.find_element_by_xpath("(//input[@class='ng-pristine ng-untouched ng-valid hasDatepicker ng-not-empty ng-valid-required'])[1]")
date2.clear()
date2.send_keys(date_ranges[1])
date2.send_keys(Keys.ESCAPE)
driver.implicitly_wait(10)

run_report_button = driver.find_element_by_css_selector("button.btn.btn-success[ng-click='runReport()']").click()
driver.implicitly_wait(80)
export_excel = driver.find_element_by_css_selector("button.ui-button.ui-widget.ui-state-default.ui-corner-all.ui-button-text-only[ng-click='exportToExcel()']").click()
time.sleep(5)
driver.quit()



######## load file and create logic + dataframe
import pandas as pd
from datetime import datetime, timedelta
import math
import glob

filename = glob.glob("..\..\..\input\care_logs*.xlsx")[0]
df = pd.read_excel(filename)

# convert from timestamp to datetime
# Early is denoted as negative
# Late is denoted as positive
def clock_in():
    for i in range(0,df.shape[0]):
        tdelta = df['Scheduled Clock In'].loc[i].to_pydatetime() - df['Actual Clock In'].loc[i].to_pydatetime()
        if tdelta.days < 0:
            new_minutes = round(1440-(tdelta.seconds/60))
            df.loc[i, 'Clock IN Difference'] = float('{}'. format(new_minutes))
        elif tdelta.days >= 0:
            new_minutes = round(tdelta.seconds/60)
            df.loc[i, 'Clock IN Difference'] = float('-{}'. format(new_minutes))


def clock_out():
    for i in range(0,df.shape[0]):
        tdelta = df['Scheduled Clock Out'].loc[i].to_pydatetime() - df['Actual Clock Out'].loc[i].to_pydatetime()
        if tdelta.days < 0:
            new_minutes = round(1440-(tdelta.seconds/60))
            df.loc[i, 'Clock OUT Difference'] = float('{}'. format(new_minutes))
        elif tdelta.days >= 0:
            new_minutes = round(tdelta.seconds/60)
            df.loc[i, 'Clock OUT Difference'] = float('-{}'. format(new_minutes))


clock_in()
clock_out()
df.fillna(0, inplace=True)

# take time input from time.txt
time_file = open('../../../time.txt', 'r',)
time_list = []
for i in time_file.read().splitlines():
    time_list.append(int(i))

time_file.close()

clc_in_early = time_list[0]
clc_in_late = time_list[1]
clc_out_early = time_list[2]
clc_out_late = time_list[3]

# query out caregivers without clock infractions
ds = df[(df['Clock IN Difference'] <= -clc_in_early) | (df['Clock IN Difference'] >= clc_in_late) |
        (df['Clock OUT Difference'] <= -clc_out_early) | (df['Clock OUT Difference'] >= clc_out_late)
       ]
ds = ds.sort_values(by='Caregiver Name')

most_infractions_df = ds.groupby(by='Caregiver Name').count()['Scheduled Clock In']
most_infractions = most_infractions_df.sort_values(ascending=False)

# find the caregivers with infractions in dataframe of all caregivers to give count of total visits of caregivers
# with infractions
shifts = df.groupby(by='Caregiver Name').count()['Scheduled Clock In']
matching_most_shifts = shifts[shifts.index.isin(most_infractions_df.index)]

# infractions per visit
most_infractions_per_shift = most_infractions/matching_most_shifts
most_infractions_per_shift_sorted = most_infractions_per_shift.sort_values(ascending=False)






# excel portion
from openpyxl.styles import PatternFill, colors, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

workbook = Workbook()
ws = workbook.active

for r in dataframe_to_rows(ds, index=False, header=True):
    ws.append(r)

# Caregivers with the most infractions
ws['L4'] = 'Most Infractions'
ws['L4'].font = Font(bold=True)
try:
    ws['L5'] = most_infractions.index[0]
except:
    ws['L5'] = ''
try:
    ws['L6'] = most_infractions.index[1]
except:
    ws['L6'] = ''
try:
    ws['L7'] = most_infractions.index[2]
except:
    ws['L7'] = ''
try:
    ws['L8'] = most_infractions.index[3]
except:
    ws['L8'] = ''
try:
    ws['L9'] = most_infractions.index[4]
except:
    ws['L9'] = ''
try:
    ws['M5'] = most_infractions[0]
except:
    ws['M5'] = ''
try:
    ws['M6'] = most_infractions[1]
except:
    ws['M6'] = ''
try:
    ws['M7'] = most_infractions[2]
except:
    ws['M7'] = ''
try:
    ws['M8'] = most_infractions[3]
except:
    ws['M8'] = ''
try:
    ws['M9'] = most_infractions[4]
except:
    ws['M9'] = ''

# Caregivers with the most infractions per visit
ws['O4'] = 'Infraction Percentage'
ws['O4'].font = Font(bold=True)
try:
    ws['O5'] = most_infractions_per_shift_sorted.index[0]
except:
    ws['O5'] = ''
try:
    ws['O6'] = most_infractions_per_shift_sorted.index[1]
except:
    ws['O6'] = ''
try:
    ws['O7'] = most_infractions_per_shift_sorted.index[2]
except:
    ws['O7'] = ''
try:
    ws['O8'] = most_infractions_per_shift_sorted.index[3]
except:
    ws['O8'] = ''
try:
    ws['O9'] = most_infractions_per_shift_sorted.index[4]
except:
    ws['O9'] = ''
try:
    ws['P5'] = most_infractions_per_shift_sorted[0]*100
except:
    ws['P5'] = ''
try:
    ws['P6'] = most_infractions_per_shift_sorted[1]*100
except:
    ws['P6'] = ''
try:
    ws['P7'] = most_infractions_per_shift_sorted[2]*100
except:
    ws['P7'] =''
try:
    ws['P8'] = most_infractions_per_shift_sorted[3]*100
except:
    ws['P8'] = ''
try:
    ws['P9'] = most_infractions_per_shift_sorted[4]*100
except:
    ws['P9'] = ''


# Caregivers with the most shifts
most_shifts = shifts.sort_values(ascending=False)
ws['R4'] = 'Most Shifts'
ws['R4'].font = Font(bold=True)
try:
    ws['R5'] = most_shifts.index[0]
except:
    ws['R5'] = ''
try:
    ws['R6'] = most_shifts.index[1]
except:
    ws['R6'] = ''
try:
    ws['R7'] = most_shifts.index[2]
except:
    ws['R7'] = ''
try:
    ws['R8'] = most_shifts.index[3]
except:
    ws['R8'] = ''
try:
    ws['R9'] = most_shifts.index[4]
except:
    ws['R9'] = ''
try:
    ws['S5'] = most_shifts[0]
except:
    ws['S5'] = ''
try:
    ws['S6'] = most_shifts[1]
except:
    ws['S6'] = ''
try:
    ws['S7'] = most_shifts[2]
except:
    ws['S7'] = ''
try:
    ws['S8'] = most_shifts[3]
except:
    ws['S8'] = ''
try:
    ws['S9'] = most_shifts[4]
except:
    ws['S9'] = ''

# Caregivers with the most late clock ins
late_clc_in_df = ds[ds['Clock IN Difference']>=clc_in_late]
cg_clc_in_late = late_clc_in_df.groupby(by='Caregiver Name').count().sort_values(by='Scheduled Clock In', ascending=False)['Scheduled Clock In']
ws['I12'] = 'Most Late Clock Ins'
ws['I12'].font = Font(bold=True)
try:
    ws['I13'] = cg_clc_in_late.index[0]
except:
    ws['I13'] = ''
try:
    ws['I14'] = cg_clc_in_late.index[1]
except:
    ws['I14'] = ''
try:
    ws['I15'] = cg_clc_in_late.index[2]
except:
    ws['I15'] = ''
try:
    ws['I16'] = cg_clc_in_late.index[3]
except:
    ws['I16'] = ''
try:
    ws['I17'] = cg_clc_in_late.index[4]
except:
    ws['I17'] = ''
try:
    ws['J13'] = cg_clc_in_late[0]
except:
    ws['J13'] = ''
try:
    ws['J14'] = cg_clc_in_late[1]
except:
    ws['J14'] = ''
try:
    ws['J15'] = cg_clc_in_late[2]
except:
    ws['J15'] = ''
try:
    ws['J16'] = cg_clc_in_late[3]
except:
    ws['J16'] = ''
try:
    ws['J17'] = cg_clc_in_late[4]
except:
    ws['J17'] = ''

# Caregivers with the most early clock outs
early_clock_out_df = ds[ds['Clock OUT Difference'] <= -clc_out_early]
cg_clc_out_early = early_clock_out_df.groupby(by='Caregiver Name').count().sort_values(by='Scheduled Clock In', ascending=False)['Scheduled Clock In']
ws['L12'] = 'Most Early Clock Outs'
ws['L12'].font = Font(bold=True)
try:
    ws['L13'] = cg_clc_out_early.index[0]
except:
    ws['L13'] = ''
try:
    ws['L14'] = cg_clc_out_early.index[1]
except:
    ws['L14'] = ''
try:
    ws['L15'] = cg_clc_out_early.index[2]
except:
    ws['L15'] = ''
try:
    ws['L16'] = cg_clc_out_early.index[3]
except:
    ws['L16'] = ''
try:
    ws['L17'] = cg_clc_out_early.index[4]
except:
    ws['L17'] = ''
try:
    ws['M13'] = cg_clc_out_early[0]
except:
    ws['M13'] = ''
try:
    ws['M14'] = cg_clc_out_early[1]
except:
    ws['M14'] = ''
try:
    ws['M15'] = cg_clc_out_early[2]
except:
    ws['M15'] = ''
try:
    ws['M16'] = cg_clc_out_early[3]
except:
    ws['M16'] = ''
try:
    ws['M17'] = cg_clc_out_early[4]
except:
    ws['M17'] = ''


# Caregivers with the most late clock outs
late_clock_out_df = ds[ds['Clock OUT Difference'] >= clc_out_late]
cg_clc_out_late = late_clock_out_df.groupby(by='Caregiver Name').count().sort_values(by='Scheduled Clock In', ascending=False)['Scheduled Clock In']
ws['O12'] = 'Most Late Clock Outs'
ws['O12'].font = Font(bold=True)
try:
    ws['O13'] = cg_clc_out_late.index[0]
except:
    ws['O13'] = ''
try:
    ws['O14'] = cg_clc_out_late.index[1]
except:
    ws['O14'] = ''
try:
    ws['O15'] = cg_clc_out_late.index[2]
except:
    ws['O15'] = ''
try:
    ws['O16'] = cg_clc_out_late.index[3]
except:
    ws['O16'] = ''
try:
    ws['O17'] = cg_clc_out_late.index[4]
except:
    ws['O17'] = ''
try:
    ws['P13'] = cg_clc_out_late[0]
except:
    ws['P13'] = ''
try:
    ws['P14'] = cg_clc_out_late[1]
except:
    ws['P14'] = ''
try:
    ws['P15'] = cg_clc_out_late[2]
except:
    ws['P15'] = ''
try:
    ws['P16'] = cg_clc_out_late[3]
except:
    ws['P16'] = ''
try:
    ws['P17'] = cg_clc_out_late[4]
except:
    ws['P17'] = ''


# Legend
ws['J4'] = 'Legend (mins)'
ws['J4'].font = Font(bold=True)
ws['J5'].fill = PatternFill(fgColor="FFC7CE", fill_type = "solid")
ws['J5'] = 'Late'
ws['J5'].font = Font(bold=True, color = 'C00000')
ws['J6'].fill = PatternFill(fgColor="000000", fill_type = "solid")
ws['J6'] = 'Early'
ws['J6'].font =Font(bold=True, color='FFFFFF')

# Clock In Early
diff_style = DifferentialStyle(font=Font(color='FFFFFF', bold=True),fill=PatternFill(bgColor="1C1F1D"))
rule1 = Rule(type="expression", dxf=diff_style)
rule1.formula = ["$F2<=-{}". format(clc_in_early)]
ws.conditional_formatting.add("F2:F5000", rule1)

# Clock In Late
diff_style = DifferentialStyle(font=Font(color='C00000', bold=True),fill=PatternFill(bgColor="FFC7CE"))
rule2 = Rule(type="expression", dxf=diff_style)
rule2.formula = ["$F2>={}". format(clc_in_late)]
ws.conditional_formatting.add("F2:F5000", rule2)

# Clock Out Early
diff_style = DifferentialStyle(font=Font(color='FFFFFF', bold=True), fill=PatternFill(bgColor='1C1F1D'))
rule3 = Rule(type="expression", dxf=diff_style)
rule3.formula = ["$G2<=-{}". format(clc_out_early)]
ws.conditional_formatting.add("G2:G5000", rule3)

# Clock Out Late
diff_style = DifferentialStyle(font=Font(color='C00000', bold=True), fill=PatternFill(bgColor='FFC7CE'))
rule4 = Rule(type="expression", dxf=diff_style)
rule4.formula = ["$G2>={}". format(clc_out_late)]
ws.conditional_formatting.add("G2:G5000", rule4)
workbook.save('..\..\..\CLOCK_infraction_report.xlsx')

for excel_file in glob.iglob(os.path.join(download_PATH, '*.xlsx')):
    os.remove(excel_file)

print("Process Complete")






















###
